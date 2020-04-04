from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from time import sleep, strftime
import openpyxl, requests, bs4, datetime
#List of Links to iterate over
urls = ['https://www.instagram.com/ucdafc/','https://www.instagram.com/ucdamericanfootball/', 'https://www.instagram.com/ucdarchery/', 'https://www.instagram.com/ucdathletics/', 'https://www.instagram.com/ucd_badminton_official/',
'https://www.instagram.com/ucdladiesbc/', 'https://www.instagram.com/ucdmensbk/', 'https://www.instagram.com/ucdbc/','https://www.instagram.com/ucdlbc/', 'https://www.instagram.com/ucd.boxing/',
'https://www.instagram.com/ucdcanoeclub/', 'https://www.instagram.com/ucdcycling/', 'https://www.instagram.com/ucdequestrian/', 'https://www.instagram.com/ucdfencing/', 'https://www.instagram.com/ucdmenshockey/', 'https://www.instagram.com/ucdladieshc/', 'https://www.instagram.com/ucdjudoclub/',
'https://www.instagram.com/ucdkarate/', 'https://www.instagram.com/ucd_kite_wake/', 'https://www.instagram.com/ucd_lacrosse/', 'https://www.instagram.com/ucdmountaineering/',
'https://www.instagram.com/ucdnetball/', 'https://www.instagram.com/orienteeringucd/', 'https://www.instagram.com/ucdrugby/', 'https://www.instagram.com/ucdwomensrugby/',
'https://www.instagram.com/ucdsailing/', 'https://www.instagram.com/ucdsepaktakraw/', 'https://www.instagram.com/ucdsnowsports/', 'https://www.instagram.com/ucdsport/',
'https://www.instagram.com/ucdsquash/', 'https://www.instagram.com/ucdsubaqua/', 'https://www.instagram.com/ucd_surfclub/', 'https://www.instagram.com/ucdswimming/',
'https://www.instagram.com/ucdtabletennis/', 'https://www.instagram.com/ucdtagrugby/', 'https://www.instagram.com/ucdtaekwondo/', 'https://www.instagram.com/ucdtennis/',
'https://www.instagram.com/ucdtrampoline/', 'https://www.instagram.com/ucdtriathlon/', 'https://www.instagram.com/ucd_ultimate/', 'https://www.instagram.com/ucdwindsurfing/']

#Path to webdriver
chromedriver_path = 'C:\\Users\\callum\\Desktop\\chromedriver_win32\\chromedriver.exe' # Change this to your own chromedriver path!

#Open Instagram and login, click not now on notification popup
webdriver = webdriver.Chrome(executable_path=chromedriver_path)
sleep(1)
webdriver.get('https://www.instagram.com/accounts/login/?source=auth_switcher')
sleep(1)

username = webdriver.find_element_by_name('username')
username.send_keys(email)#Your instagram account email
password = webdriver.find_element_by_name('password')
password.send_keys(password)#Your instagram account password

button_login = webdriver.find_element_by_css_selector('#react-root > section > main > div > article > div > div:nth-child(1) > div > form > div:nth-child(4) > button')
button_login.click()
sleep(3)


notnow = webdriver.find_element_by_css_selector('body > div.RnEpo.Yx5HN > div > div > div.mt3GC > button.aOOlW.HoLwm')
notnow.click()

#Open the workbook 
wb = openpyxl.load_workbook('ClubFollow.xlsx')

sheet = wb['Sheet1']

col = 2
#find the first empty column
while sheet.cell(row=4, column=col).value != None:
    col += 1
    
#put todays date on the first row
sheet.cell(row=1, column=col).value = str(datetime.datetime.now().date())
def getData(url, i):
    """
    Loop through all the urls, and find the follower count based on XPath.
    Then place the value in the correct workbook cell

    """
    webdriver.get(url)
    # For init setup of workbook
    # soup = bs4.BeautifulSoup(webdriver.page_source, 'html.parser')
    # data = soup.find_all('meta', attrs={'property': 'og:description'})
    # text = data[0].get('content').split()
    # user = '%s %s %s' % (text[-3], text[-2], text[-1])
    num = webdriver.find_element_by_xpath('/html/body/div[1]/section/main/div/header/section/ul/li[2]/a/span').text
    sheet.cell(row=i, column=col).value = int(num.replace(',',''))
    sleep(1)

for i in range(len(urls)):
    getData(urls[i], i+2)



wb.save('ClubFollow.xlsx')
webdriver.close()